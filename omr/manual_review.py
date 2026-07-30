"""
Manual roll-number review.

A sheet lands here when its bubbled roll number can't be trusted
automatically - either a digit was left blank/double-marked, or the
digits read cleanly but don't match any real student in the school's
own records (the API lookup came back empty). Rather than guessing
(e.g. from a filename), the sheet is set aside with a cropped image of
its roll-number grid so a person can read it by eye and type in the
correct roll number.

Produces manual_review.xlsx: one row per flagged sheet, with the
cropped roll-grid image embedded next to a blank "Actual Roll No"
column. After a person fills that column in and saves the file, run
tools/resolve_manual_roll.py to finish those sheets (API lookup +
final PDF) without re-scanning anything.
"""

import os

import cv2
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

CROP_PADDING = 25
IMG_DISPLAY_WIDTH_PX = 260


def crop_roll_region(warped_bgr, layout, padding=CROP_PADDING):
    """Crop the roll-number bubble grid out of a warped sheet image,
    for a human to visually read."""
    roll_grid = layout.get("roll_grid")
    if not roll_grid:
        return warped_bgr

    xs, ys = [], []
    for values in roll_grid["bubbles"].values():
        for x, y in values.values():
            xs.append(x)
            ys.append(y)

    radius = layout.get("bubble_radius", 20)
    x0 = max(0, int(min(xs) - radius - padding))
    x1 = min(warped_bgr.shape[1], int(max(xs) + radius + padding))
    y0 = max(0, int(min(ys) - radius - padding * 3))  # extra top margin for the digit boxes above the bubbles
    y1 = min(warped_bgr.shape[0], int(max(ys) + radius + padding))

    return warped_bgr[y0:y1, x0:x1]


def save_roll_crop(warped_bgr, layout, out_dir, base_name):
    os.makedirs(out_dir, exist_ok=True)
    crop = crop_roll_region(warped_bgr, layout)
    out_path = os.path.join(out_dir, f"{base_name}_roll.png")
    cv2.imwrite(out_path, crop)
    return out_path


def build_manual_review_workbook(pending_entries, out_path):
    """
    pending_entries: list of dicts with keys:
        filename, image_path, detected_roll (str or None), reason (str)
    Writes a standalone manual_review.xlsx. Returns out_path, or None if
    there was nothing to write.
    """
    if not pending_entries:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manual Review"

    headers = ["Filename", "Roll Number Grid (from scan)", "Bubbles Read As", "Reason", "Actual Roll No"]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 16

    for r, entry in enumerate(pending_entries, start=2):
        ws.cell(row=r, column=1, value=entry["filename"])
        ws.cell(row=r, column=3, value=entry.get("detected_roll") or "(unclear)")
        ws.cell(row=r, column=4, value=entry["reason"])
        ws.row_dimensions[r].height = 90

        if entry.get("image_path") and os.path.exists(entry["image_path"]):
            img = XLImage(entry["image_path"])
            aspect = img.height / img.width
            img.width = IMG_DISPLAY_WIDTH_PX
            img.height = int(IMG_DISPLAY_WIDTH_PX * aspect)
            ws.add_image(img, f"B{r}")

    wb.save(out_path)
    return out_path
