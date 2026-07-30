"""
Writes the final Excel report: one row per student in "Results", a
per-question breakdown in "Question Analysis", and a full audit trail
in "Answer Detail" (every detected answer per student, for manual spot
checks / dispute resolution).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from omr.answer_key import format_clauses

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _write_header(ws, headers):
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fmt_answer(ans_tuple):
    return " ".join(ans_tuple) if ans_tuple else "-"


def build_report(output_path, student_results, question_stats, warnings=None):
    """
    student_results: list of dicts, each with keys:
        filename, roll_no, summary (dict from score_sheet), per_question (list),
        needs_review (bool), student_info (dict or None - id_no/name/
        father_name/class/section from the school's API, once resolved)
    question_stats: output of aggregate_question_stats()
    """
    wb = openpyxl.Workbook()

    # ---------- Results sheet ----------
    ws = wb.active
    ws.title = "Results"
    headers = [
        "Roll No", "ID No", "Name", "Father Name", "Class", "Section", "File",
        "Attempted", "Correct", "Wrong", "Unattempted", "Bonus",
        "Marks Obtained", "Max Marks", "Percentage", "Needs Manual Roll Review",
    ]
    _write_header(ws, headers)

    for r, res in enumerate(student_results, start=2):
        s = res["summary"]
        info = res.get("student_info") or {}
        row = [
            res["roll_no"] or "PENDING REVIEW",
            info.get("id_no", ""), info.get("name", ""), info.get("father_name", ""),
            info.get("class", ""), info.get("section", ""),
            res["filename"],
            s["attempted"], s["correct"], s["wrong"], s["unattempted"], s.get("bonus", 0),
            s["marks_obtained"], s["max_marks"], s["percentage"],
            "YES" if res["needs_review"] else "",
        ]
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if res["needs_review"]:
                cell.fill = FLAG_FILL

    _autosize(ws, [14, 10, 22, 22, 8, 9, 26, 10, 9, 9, 11, 8, 14, 10, 11, 20])

    # ---------- Question Analysis sheet ----------
    ws2 = wb.create_sheet("Question Analysis")
    headers2 = ["Question", "Correct Answer", "Marks", "Negative Marks", "Bonus", "Multi Correct",
                "A", "B", "C", "D", "Unattempted", "Correct Count", "Facility Index %"]
    _write_header(ws2, headers2)
    for r, q in enumerate(question_stats, start=2):
        row = [
            q["question"], format_clauses(q["correct_answer"]), q["marks"], q["negative_marks"],
            "YES" if q["is_bonus"] else "", "YES" if q["is_multi_correct"] else "",
            q["A"], q["B"], q["C"], q["D"],
            q["unattempted"], q["correct_count"], q["facility_index_%"],
        ]
        for c, val in enumerate(row, start=1):
            ws2.cell(row=r, column=c, value=val)
    _autosize(ws2, [10, 16, 8, 12, 8, 11, 6, 6, 6, 6, 12, 13, 16])

    # ---------- Answer Detail sheet (audit trail) ----------
    # Columns are built from whichever question numbers actually appear
    # (i.e. only questions the answer key covers) rather than assuming a
    # fixed 1..100 range, so an exam with fewer than 100 questions simply
    # doesn't get columns for the rest. Multi-answer questions show as
    # e.g. "B C"; a blank is shown as "-". A single-answer question with
    # an accidental double-mark that got resolved to its most confident
    # option (see omr/scoring.py) shows as e.g. "B (of B C)" so the raw
    # shaded bubbles are still visible for a spot check.
    ws3 = wb.create_sheet("Answer Detail")
    active_qnums = sorted({pq["question"] for res in student_results for pq in res["per_question"]})
    qnum_to_col = {q: 3 + i for i, q in enumerate(active_qnums)}
    detail_headers = ["Roll No", "File"] + [f"Q{q}" for q in active_qnums]
    _write_header(ws3, detail_headers)
    for r, res in enumerate(student_results, start=2):
        ws3.cell(row=r, column=1, value=res["roll_no"] or "PENDING REVIEW")
        ws3.cell(row=r, column=2, value=res["filename"])
        for pq in res["per_question"]:
            col = qnum_to_col[pq["question"]]
            if pq["verdict"] == "bonus":
                display = "BONUS"
            elif pq.get("resolved"):
                display = f"{_fmt_answer(pq['student_answer'])} (of {_fmt_answer(pq['raw_student_answer'])})"
            else:
                display = _fmt_answer(pq["student_answer"])
            cell = ws3.cell(row=r, column=col, value=display)
            if pq["verdict"] == "wrong":
                cell.font = Font(color="C00000")
            elif pq["verdict"] == "bonus":
                cell.font = Font(color="1F4E78", italic=True)
            elif pq.get("resolved"):
                cell.font = Font(color="9C5700", italic=True)
    _autosize(ws3, [14, 26] + [6] * len(active_qnums))

    # ---------- Warnings sheet ----------
    if warnings:
        ws4 = wb.create_sheet("Warnings")
        _write_header(ws4, ["Warning"])
        for r, w in enumerate(warnings, start=2):
            ws4.cell(row=r, column=1, value=w)
        _autosize(ws4, [100])

    wb.save(output_path)
