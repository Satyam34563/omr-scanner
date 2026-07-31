"""
The OMR batch-processing pipeline as a plain importable function, shared
by both the CLI (main.py) and the web API (api.py) - so both entry
points run exactly the same logic instead of two copies drifting apart.

run_batch() does everything main.py used to do inline: rasterizes the
scan PDF (or reads a folder of images), perspective-corrects and scores
every sheet, looks up each validated student against the school's
records, and writes the Excel report, combined result PDF, bubble
overlay PDF, and manual-review workbook. It raises a plain exception on
a hard failure (missing layout.json, no scans found, nothing could be
processed) instead of calling sys.exit(), since a library function
shouldn't kill the whole process - the caller (CLI prints + exits; API
turns it into an HTTP error) decides what "failure" means for it.
"""

import glob
import json
import os
from pathlib import Path

import cv2
import openpyxl

from omr.pdf_input import rasterize_scan_pdf
from omr.preprocessing import warp_to_canonical
from omr.bubble_detector import detect_answers, detect_roll_number
from omr.answer_key import load_answer_key, parse_answer_expression
from omr.scoring import score_sheet, aggregate_question_stats, _is_correct
from omr.report import build_report
from omr.student_report import build_combined_report
from omr.student_info import fetch_student_info, fetch_student_photo
from omr.manual_review import save_roll_crop, build_manual_review_workbook
from omr.overlay import build_overlay_pdf

IMAGE_EXTENSIONS = ("*.jpg", "*.jpeg", "*.png", "*.tif", "*.tiff", "*.bmp")


def find_scan_files(scans_dir):
    files = []
    for ext in IMAGE_EXTENSIONS:
        files.extend(glob.glob(os.path.join(scans_dir, ext)))
        files.extend(glob.glob(os.path.join(scans_dir, ext.upper())))
    return sorted(set(files))


def load_scan_pages(scans_pdf_path, scans_dir_path, config):
    """Returns a list of (filename, image_bgr) tuples, one per sheet to
    process, in a stable order. Prefers a single combined PDF
    (one page per sheet - the primary, scanner-friendly workflow) and
    falls back to a folder of separate image files only if that PDF
    doesn't exist. Raises FileNotFoundError if neither is available."""
    if scans_pdf_path and os.path.exists(scans_pdf_path):
        pages = rasterize_scan_pdf(scans_pdf_path, dpi=config.get("scan_pdf_dpi", 200))
        stem = Path(scans_pdf_path).stem
        width = max(3, len(str(len(pages))))
        return [(f"{stem}_page{i + 1:0{width}d}", img) for i, img in enumerate(pages)]

    scan_files = find_scan_files(scans_dir_path) if scans_dir_path else []
    if not scan_files:
        raise FileNotFoundError(
            f"No scans found. Put a combined PDF at '{scans_pdf_path}' (one page per sheet), "
            f"or image files in '{scans_dir_path}'."
        )
    return [(os.path.basename(p), cv2.imread(p)) for p in scan_files]


def run_batch(
    scans_pdf_path,
    scans_dir_path,
    answer_key_path,
    config,
    layout,
    output_path,
    reports_pdf_path,
    manual_review_path,
    overlay_pdf_path,
    progress_callback=None,
):
    """
    Runs the full batch: scan -> detect -> score -> student lookup ->
    write every output file. `config` and `layout` are already-loaded
    dicts (not paths), so a caller that keeps them in memory across
    requests (like the web API) doesn't have to re-read them from disk
    every time. `progress_callback(filename, index, total, confirmed_roll,
    summary)`, if given, is called once per sheet as it finishes -
    useful for printing live progress (CLI) or reporting batch progress
    (web API). `confirmed_roll` and `summary` are both None if the
    sheet failed to process at all (e.g. couldn't be warped).

    Returns a dict:
        {
            "output_path": str,
            "reports_pdf_path": str or None,   # None if nothing was resolved yet
            "reports_pdf_manifest": [{"roll_no": str, "page": int}, ...],  # 1-indexed page order
            "overlay_pdf_path": str or None,
            "manual_review_path": str or None, # None if nothing needs review
            "num_processed": int,
            "num_resolved": int,
            "num_pending_review": int,
            "warnings": [str, ...],
        }
    Raises RuntimeError if no sheets could be processed at all.
    """
    target_markers = layout.get("markers")

    answer_key, key_warnings = load_answer_key(
        answer_key_path, num_questions=config["num_questions"], marking_scheme=config["marking_scheme"]
    )
    warnings = list(key_warnings)

    scan_pages = load_scan_pages(scans_pdf_path, scans_dir_path, config)

    student_results = []
    all_per_question = []
    pending_entries = []
    overlay_pages = []
    pending_review_dir = config.get("pending_review_dir", "output/pending_review")

    for i, (filename, raw_image) in enumerate(scan_pages):
        try:
            if raw_image is None:
                raise ValueError("could not read page/image data")
            warped = warp_to_canonical(
                raw_image, layout["canonical_width"], layout["canonical_height"], target_markers=target_markers
            )
        except Exception as e:
            warnings.append(f"Failed to process '{filename}': {e}")
            if progress_callback:
                progress_callback(filename, i + 1, len(scan_pages), None, None)
            continue

        # --- Roll number: read from the bubbled grid, then VALIDATE
        # against the school's own student database. Either step
        # failing routes this sheet to manual review instead of
        # guessing - see omr/manual_review.py.
        bubbled_roll, roll_digits, roll_unclear = detect_roll_number(
            warped, layout, fill_threshold=config["fill_threshold"], ambiguous_margin=config["ambiguous_margin"]
        )

        student_info = None
        reason = None
        if roll_unclear or bubbled_roll is None:
            reason = "One or more roll-number digits could not be read clearly."
        else:
            student_info = fetch_student_info(bubbled_roll, config)
            if student_info is None:
                reason = f"Bubbled roll number ({bubbled_roll}) was not found in student records."
            else:
                photo_dir = config.get("student_photo_dir", "output/student_photos")
                student_info["photo_path"] = fetch_student_photo(
                    student_info.get("image_url"), photo_dir, student_info["id_no"]
                )
                if student_info.get("image_url") and not student_info["photo_path"]:
                    warnings.append(
                        f"'{filename}': student photo could not be downloaded from "
                        f"{student_info['image_url']} (network issue or file missing) - "
                        f"PDF will show a 'No Photo' box instead."
                    )

        confirmed_roll = bubbled_roll if (student_info is not None) else None

        if confirmed_roll is None:
            crop_path = save_roll_crop(warped, layout, pending_review_dir, os.path.splitext(filename)[0])
            pending_entries.append({
                "filename": filename,
                "image_path": crop_path,
                "detected_roll": bubbled_roll,
                "reason": reason,
            })
            warnings.append(f"'{filename}': {reason} Sent to manual review.")

        detected = detect_answers(
            warped, layout,
            fill_threshold=config["fill_threshold"],
            ambiguous_margin=config["ambiguous_margin"],
        )
        # 3 or 4 options "filled" on one question is essentially never a
        # genuine answer - it's almost always a scribbled-out/changed
        # answer, a crease, a stain, or heavy ink from a neighboring row
        # bleeding across several bubbles at once. Score it as detected
        # (so it isn't silently dropped), but flag it - it's exactly the
        # kind of question worth a human glancing at the physical sheet for.
        suspect_questions = [q for q, d in detected.items() if len(d["answers"]) >= 3]
        if suspect_questions:
            warnings.append(
                f"'{filename}': question(s) {', '.join(str(q) for q in sorted(suspect_questions))} "
                f"had 3+ bubbles detected as filled - likely a stray mark, crease, or ink bleeding "
                f"in from a neighboring row rather than a genuine answer. Worth checking against the "
                f"physical sheet."
            )

        summary, per_question = score_sheet(detected, answer_key, config["marking_scheme"])
        needs_review = confirmed_roll is None

        # A single-answer question with 2+ bubbles shaded gets resolved
        # to the most confidently-filled one for scoring (see
        # omr/scoring.py) rather than failed outright - worth a note
        # since it means the physical sheet had an accidental double-mark.
        resolved_entries = [pq for pq in per_question if pq.get("resolved")]
        if resolved_entries:
            details = ", ".join(
                f"Q{pq['question']} ({' '.join(pq['raw_student_answer'])} -> {pq['student_answer'][0]})"
                for pq in resolved_entries
            )
            warnings.append(
                f"'{filename}': {len(resolved_entries)} single-answer question(s) had more than one "
                f"bubble shaded - scored using the most confidently-filled option: {details}."
            )

        overlay_pages.append((filename, warped, detected, roll_digits, confirmed_roll or bubbled_roll))

        student_results.append({
            "filename": filename,
            "roll_no": confirmed_roll,
            "summary": summary,
            "per_question": per_question,
            "needs_review": needs_review,
            "student_info": student_info,
        })
        all_per_question.append(per_question)

        if progress_callback:
            progress_callback(filename, i + 1, len(scan_pages), confirmed_roll, summary)

    if not student_results:
        raise RuntimeError("No sheets could be processed.")

    question_stats = aggregate_question_stats(all_per_question, answer_key)
    build_report(output_path, student_results, question_stats, warnings)

    combined_path, skipped, reports_pdf_manifest = build_combined_report(student_results, config, reports_pdf_path)
    overlay_path = build_overlay_pdf(overlay_pages, layout, overlay_pdf_path)
    manual_review_path_out = build_manual_review_workbook(pending_entries, manual_review_path)

    return {
        "output_path": output_path,
        "reports_pdf_path": combined_path,
        "reports_pdf_manifest": reports_pdf_manifest,
        "overlay_pdf_path": overlay_path,
        "manual_review_path": manual_review_path_out,
        "num_processed": len(student_results),
        "num_resolved": len(student_results) - len(skipped),
        "num_pending_review": len(pending_entries),
        # JSON-safe view of pending_entries for a web UI to render a
        # review table from - image_filename matches whatever the
        # caller zips/serves the crop images as (see api.py), not a
        # full filesystem path (which wouldn't mean anything to a
        # client anyway).
        "pending_entries": [
            {
                "filename": e["filename"],
                "detected_roll": e["detected_roll"],
                "reason": e["reason"],
                "image_filename": os.path.basename(e["image_path"]) if e.get("image_path") else None,
            }
            for e in pending_entries
        ],
        "warnings": warnings,
    }


def load_config_and_layout(config_path, layout_path):
    """Small shared helper: load config.json and layout.json from disk,
    raising a clear error if layout.json hasn't been generated yet."""
    with open(config_path) as f:
        config = json.load(f)
    if not os.path.exists(layout_path):
        raise FileNotFoundError(
            f"Layout file '{layout_path}' not found. Generate it automatically from the master PDF:\n"
            f"  python tools/auto_generate_layout.py --pdf {config.get('sheet_pdf', '<sheet>.pdf')} --output {layout_path}"
        )
    with open(layout_path) as f:
        layout = json.load(f)
    return config, layout


# --- Manual review resolution -----------------------------------------
#
# A sheet whose roll number couldn't be read/matched during run_batch()
# gets set aside rather than guessed at (see omr/manual_review.py).
# resolve_manual_review() finishes those sheets once a person has
# figured out the correct roll number - by re-scoring against the
# bubbles already detected and saved in results.xlsx's "Answer Detail"
# sheet, so none of the original scan images are needed again. It's
# shared by the CLI (tools/resolve_manual_roll.py, which parses
# corrections out of a filled-in manual_review.xlsx) and the web API
# (api.py's POST /resolve, which takes corrections as plain JSON from
# a web form instead).

def _read_answer_key_from_report(results_path, marking_scheme):
    """Rebuilds the same {q_num: {"clauses":..., "marks":..., ...}}
    structure omr.answer_key.load_answer_key() produces, but from the
    already-saved Question Analysis sheet (which has the formatted
    "A or B" / "A and B" style Correct Answer text plus this exam's
    actual per-question Marks/Negative Marks/Bonus/Multi Correct) -
    rather than re-reading the original answer key file - so a
    question's marking stays exactly as it was for the original run
    even if the answer key file has since been edited."""
    wb = openpyxl.load_workbook(results_path, data_only=True)
    ws = wb["Question Analysis"]
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    q_col = header.index("Question")
    ans_col = header.index("Correct Answer")
    marks_col = header.index("Marks") if "Marks" in header else None
    neg_col = header.index("Negative Marks") if "Negative Marks" in header else None
    bonus_col = header.index("Bonus") if "Bonus" in header else None
    multi_col = header.index("Multi Correct") if "Multi Correct" in header else None

    default_marks = marking_scheme.get("correct", 1.0)
    default_negative_marks = abs(marking_scheme.get("wrong", 0.0))

    answer_key = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        q_num = row[q_col]
        if q_num is None:
            continue
        is_multi_correct = bool(multi_col is not None and row[multi_col] == "YES")
        clauses = parse_answer_expression(row[ans_col], is_multi_correct) or []
        is_bonus = bool(bonus_col is not None and row[bonus_col] == "YES")
        marks = row[marks_col] if marks_col is not None and row[marks_col] not in (None, "") else default_marks
        negative_marks = (
            row[neg_col] if neg_col is not None and row[neg_col] not in (None, "") else default_negative_marks
        )
        answer_key[int(q_num)] = {
            "clauses": clauses, "marks": float(marks), "negative_marks": float(negative_marks),
            "is_bonus": is_bonus, "is_multi_correct": is_multi_correct,
        }
    return answer_key


def _reconstruct_per_question(ws_detail, filename, answer_key, marking_scheme):
    header = [str(c.value).strip() if c.value else "" for c in ws_detail[1]]
    file_col = header.index("File")
    q_cols = {int(h[1:]): i for i, h in enumerate(header) if h.startswith("Q") and h[1:].isdigit()}

    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        if row[file_col] == filename:
            per_question = []
            for q_num in sorted(q_cols.keys()):
                entry = answer_key.get(q_num, {"clauses": [], "marks": marking_scheme.get("correct", 1.0),
                                                "negative_marks": abs(marking_scheme.get("wrong", 0.0)),
                                                "is_bonus": False})
                val = row[q_cols[q_num]]
                is_bonus_cell = str(val).strip().upper() == "BONUS"
                # A resolved double-mark is saved as e.g. "B (of B C)" -
                # only the part before the parenthetical is the actual
                # (already-resolved) answer used for grading.
                was_resolved = " (of " in str(val) if val else False
                primary_val = str(val).split(" (")[0].strip() if val else ""
                student_ans = tuple(primary_val.split()) if primary_val and primary_val != "-" and not is_bonus_cell else ()

                if entry["is_bonus"]:
                    verdict = "bonus"
                elif not student_ans:
                    verdict = "unattempted"
                elif _is_correct(student_ans, entry["clauses"]):
                    verdict = "correct"
                else:
                    verdict = "wrong"

                per_question.append({
                    "question": q_num, "student_answer": student_ans, "resolved": was_resolved,
                    "correct_answer": entry["clauses"], "verdict": verdict,
                    "marks": entry["marks"], "negative_marks": entry["negative_marks"],
                    "is_bonus": entry["is_bonus"],
                })
            return per_question
    return None


def _score_summary_from_questions(per_question, marking_scheme):
    correct = sum(1 for p in per_question if p["verdict"] in ("correct", "bonus"))
    wrong = sum(1 for p in per_question if p["verdict"] == "wrong")
    unattempted = sum(1 for p in per_question if p["verdict"] == "unattempted")
    bonus = sum(1 for p in per_question if p["verdict"] == "bonus")
    unattempted_marks = marking_scheme.get("unattempted", 0.0)

    marks = 0.0
    max_marks = 0.0
    for p in per_question:
        max_marks += p["marks"]
        if p["verdict"] in ("correct", "bonus"):
            marks += p["marks"]
        elif p["verdict"] == "wrong":
            marks -= p["negative_marks"]
        else:
            marks += unattempted_marks

    return {
        "correct": correct, "wrong": wrong, "unattempted": unattempted, "bonus": bonus,
        "attempted": correct + wrong - bonus, "total_questions": len(per_question),
        "marks_obtained": round(marks, 2), "max_marks": round(max_marks, 2),
        "percentage": round((marks / max_marks) * 100, 2) if max_marks else 0.0,
    }


def _find_cached_photo(photo_dir, id_no):
    """Looks for an already-downloaded photo for this student (from a
    previous run) instead of re-fetching it from the API every time
    the combined PDF is rebuilt."""
    if not id_no:
        return None
    matches = glob.glob(os.path.join(photo_dir, f"{id_no}.*"))
    return matches[0] if matches else None


def resolve_manual_review(results_path, corrections, config, reports_pdf_path):
    """
    Finishes sheets that were set aside for manual roll-number review.

    corrections: {filename: roll_number_str, ...} - typically typed in
    by a person after looking at that sheet's roll-grid crop image.
    Blank/whitespace-only values are skipped (treated as "not corrected
    this round" rather than an error), so callers don't need to filter
    partially-filled forms before calling this.

    Modifies results.xlsx IN PLACE (updates the Results and Answer
    Detail sheets for anything successfully resolved) and rebuilds the
    ONE combined result PDF from every currently-validated sheet -
    both ones resolved just now and any validated in the original run
    - since there's no per-student file to incrementally append to.

    Returns:
        {
            "resolved": [{"filename":, "roll_no":, "name":}, ...],
            "still_invalid": [{"filename":, "entered_roll":, "reason":}, ...],
            "reports_pdf_path": str or None,
            "reports_pdf_manifest": [{"roll_no": str, "page": int}, ...],  # 1-indexed page order
            "num_resolved": int,
            "num_still_pending": int,  # rows STILL "PENDING REVIEW" after this call
        }
    """
    answer_key = _read_answer_key_from_report(results_path, config["marking_scheme"])
    photo_dir = config.get("student_photo_dir", "output/student_photos")

    wb = openpyxl.load_workbook(results_path)
    ws = wb["Results"]
    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header)}

    ws_detail = wb["Answer Detail"]
    detail_header = [c.value for c in ws_detail[1]]
    detail_file_col = detail_header.index("File") + 1
    detail_roll_col = 1  # "Roll No" is always column A

    resolved, still_invalid = [], []

    for filename, actual_roll in corrections.items():
        actual_roll = str(actual_roll).strip() if actual_roll is not None else ""
        if not actual_roll:
            continue

        student_info = fetch_student_info(actual_roll, config)
        if student_info is None:
            still_invalid.append({
                "filename": filename, "entered_roll": actual_roll,
                "reason": f"Roll number '{actual_roll}' was not found in student records.",
            })
            continue

        student_info["photo_path"] = fetch_student_photo(
            student_info.get("image_url"), photo_dir, student_info["id_no"]
        )

        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=col["File"]).value == filename:
                ws.cell(row=r, column=col["Roll No"], value=actual_roll)
                ws.cell(row=r, column=col["ID No"], value=student_info["id_no"])
                ws.cell(row=r, column=col["Name"], value=student_info["name"])
                ws.cell(row=r, column=col["Father Name"], value=student_info["father_name"])
                ws.cell(row=r, column=col["Class"], value=student_info["class"])
                ws.cell(row=r, column=col["Section"], value=student_info["section"])
                ws.cell(row=r, column=col["Needs Manual Roll Review"], value="")
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = openpyxl.styles.PatternFill(fill_type=None)
                break

        for r in range(2, ws_detail.max_row + 1):
            if ws_detail.cell(row=r, column=detail_file_col).value == filename:
                ws_detail.cell(row=r, column=detail_roll_col, value=actual_roll)
                break

        resolved.append({"filename": filename, "roll_no": actual_roll, "name": student_info["name"]})

    wb.save(results_path)

    all_results = []
    num_still_pending = 0
    for r in range(2, ws.max_row + 1):
        roll_no = ws.cell(row=r, column=col["Roll No"]).value
        if not roll_no or roll_no == "PENDING REVIEW":
            num_still_pending += 1
            continue
        filename = ws.cell(row=r, column=col["File"]).value
        per_question = _reconstruct_per_question(ws_detail, filename, answer_key, config["marking_scheme"])
        if per_question is None:
            continue
        summary = _score_summary_from_questions(per_question, config["marking_scheme"])
        id_no = ws.cell(row=r, column=col["ID No"]).value
        student_info = {
            "id_no": str(id_no) if id_no else "",
            "name": ws.cell(row=r, column=col["Name"]).value or "",
            "father_name": ws.cell(row=r, column=col["Father Name"]).value or "",
            "class": ws.cell(row=r, column=col["Class"]).value or "",
            "section": ws.cell(row=r, column=col["Section"]).value or "",
            "photo_path": _find_cached_photo(photo_dir, id_no),
        }
        all_results.append({
            "filename": filename, "roll_no": str(roll_no), "summary": summary,
            "per_question": per_question, "needs_review": False, "student_info": student_info,
        })

    combined_path, _, reports_pdf_manifest = build_combined_report(all_results, config, reports_pdf_path)

    return {
        "resolved": resolved,
        "still_invalid": still_invalid,
        "reports_pdf_path": combined_path,
        "reports_pdf_manifest": reports_pdf_manifest,
        "num_resolved": len(resolved),
        "num_still_pending": num_still_pending,
    }
