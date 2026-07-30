"""
Finishes sheets that were set aside for manual roll-number review.

Workflow:
  1. main.py writes manual_review.xlsx for any sheet whose roll number
     couldn't be read cleanly or didn't match a real student. Each row
     has a cropped image of that sheet's roll-number grid.
  2. A person opens manual_review.xlsx, looks at each image, and types
     the correct roll number into the "Actual Roll No" column, then
     saves the file.
  3. Run this script. For every row with an "Actual Roll No" filled
     in, it looks up that roll number in the school's student database,
     reconstructs the sheet's scoring from results.xlsx (no re-scanning
     needed - the detected answers are already in the "Answer Detail"
     sheet), and updates their row in results.xlsx (Roll No, student
     info, Needs Manual Roll Review).

     There's only ONE result PDF for the whole batch
     (output/student_reports.pdf), so this script rebuilds it from
     scratch every time it runs - covering every already-validated
     student plus whatever got resolved just now - rather than
     appending to some per-student file that no longer exists.

Usage:
    python tools/resolve_manual_roll.py --results output/results.xlsx --manual-review output/manual_review.xlsx --config config.json --reports-pdf output/student_reports.pdf
"""

import argparse
import glob
import json
import os
import sys

import openpyxl

# Allow running as `python tools/resolve_manual_roll.py` from the project
# root: Python puts this script's own directory (tools/) on sys.path[0],
# not the project root, so the `omr` package (one level up) wouldn't be
# importable without this.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from omr.student_info import fetch_student_info, fetch_student_photo
from omr.student_report import build_combined_report
from omr.answer_key import parse_answer_expression
from omr.scoring import _is_correct


def _read_manual_review(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Manual Review"] if "Manual Review" in wb.sheetnames else wb.active
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    filename_col = header.index("Filename")
    actual_roll_col = header.index("Actual Roll No")

    resolved = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        filename = row[filename_col]
        actual_roll = row[actual_roll_col]
        if filename and actual_roll:
            resolved[str(filename)] = str(actual_roll).strip()
    return resolved


def _read_answer_key_from_report(results_path, marking_scheme):
    """Rebuilds the same {q_num: {"clauses":..., "marks":..., ...}}
    structure omr.answer_key.load_answer_key() produces, but from the
    already-saved Question Analysis sheet (which has the formatted
    "A or B" / "A and B" style Correct Answer text plus this exam's
    actual per-question Marks/Negative Marks/Bonus/Multi Correct) -
    rather than re-reading the original answer key file - so a
    question's marking stays exactly as it was for the original run
    even if the answer key file has since been edited."""
    wb = openpyxl.load_workbook(results_path, data_only=True)
    ws = wb["Question Analysis"]
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    q_col = header.index("Question")
    ans_col = header.index("Correct Answer")
    marks_col = header.index("Marks") if "Marks" in header else None
    neg_col = header.index("Negative Marks") if "Negative Marks" in header else None
    bonus_col = header.index("Bonus") if "Bonus" in header else None
    multi_col = header.index("Multi Correct") if "Multi Correct" in header else None

    default_marks = marking_scheme.get("correct", 1.0)
    default_negative_marks = abs(marking_scheme.get("wrong", 0.0))

    answer_key = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        q_num = row[q_col]
        if q_num is None:
            continue
        is_multi_correct = bool(multi_col is not None and row[multi_col] == "YES")
        clauses = parse_answer_expression(row[ans_col], is_multi_correct) or []
        is_bonus = bool(bonus_col is not None and row[bonus_col] == "YES")
        marks = row[marks_col] if marks_col is not None and row[marks_col] not in (None, "") else default_marks
        negative_marks = (
            row[neg_col] if neg_col is not None and row[neg_col] not in (None, "") else default_negative_marks
        )
        answer_key[int(q_num)] = {
            "clauses": clauses, "marks": float(marks), "negative_marks": float(negative_marks),
            "is_bonus": is_bonus, "is_multi_correct": is_multi_correct,
        }
    return answer_key


def _reconstruct_per_question(ws_detail, filename, answer_key, marking_scheme):
    header = [str(c.value).strip() if c.value else "" for c in ws_detail[1]]
    file_col = header.index("File")
    q_cols = {int(h[1:]): i for i, h in enumerate(header) if h.startswith("Q") and h[1:].isdigit()}

    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        if row[file_col] == filename:
            per_question = []
            for q_num in sorted(q_cols.keys()):
                entry = answer_key.get(q_num, {"clauses": [], "marks": marking_scheme.get("correct", 1.0),
                                                "negative_marks": abs(marking_scheme.get("wrong", 0.0)),
                                                "is_bonus": False})
                val = row[q_cols[q_num]]
                is_bonus_cell = str(val).strip().upper() == "BONUS"
                # A resolved double-mark is saved as e.g. "B (of B C)" -
                # only the part before the parenthetical is the actual
                # (already-resolved) answer used for grading.
                was_resolved = " (of " in str(val) if val else False
                primary_val = str(val).split(" (")[0].strip() if val else ""
                student_ans = tuple(primary_val.split()) if primary_val and primary_val != "-" and not is_bonus_cell else ()

                if entry["is_bonus"]:
                    verdict = "bonus"
                elif not student_ans:
                    verdict = "unattempted"
                elif _is_correct(student_ans, entry["clauses"]):
                    verdict = "correct"
                else:
                    verdict = "wrong"

                per_question.append({
                    "question": q_num, "student_answer": student_ans, "resolved": was_resolved,
                    "correct_answer": entry["clauses"], "verdict": verdict,
                    "marks": entry["marks"], "negative_marks": entry["negative_marks"],
                    "is_bonus": entry["is_bonus"],
                })
            return per_question
    return None


def _score_summary(per_question, marking_scheme):
    correct = sum(1 for p in per_question if p["verdict"] in ("correct", "bonus"))
    wrong = sum(1 for p in per_question if p["verdict"] == "wrong")
    unattempted = sum(1 for p in per_question if p["verdict"] == "unattempted")
    bonus = sum(1 for p in per_question if p["verdict"] == "bonus")
    unattempted_marks = marking_scheme.get("unattempted", 0.0)

    marks = 0.0
    max_marks = 0.0
    for p in per_question:
        max_marks += p["marks"]
        if p["verdict"] in ("correct", "bonus"):
            marks += p["marks"]
        elif p["verdict"] == "wrong":
            marks -= p["negative_marks"]
        else:
            marks += unattempted_marks

    return {
        "correct": correct, "wrong": wrong, "unattempted": unattempted, "bonus": bonus,
        "attempted": correct + wrong - bonus, "total_questions": len(per_question),
        "marks_obtained": round(marks, 2), "max_marks": round(max_marks, 2),
        "percentage": round((marks / max_marks) * 100, 2) if max_marks else 0.0,
    }


def _find_cached_photo(photo_dir, id_no):
    """Looks for an already-downloaded photo for this student (from a
    previous run) instead of re-fetching it from the API every time
    the combined PDF is rebuilt."""
    if not id_no:
        return None
    matches = glob.glob(os.path.join(photo_dir, f"{id_no}.*"))
    return matches[0] if matches else None


def main():
    parser = argparse.ArgumentParser(description="Resolve manually-reviewed roll numbers")
    parser.add_argument("--results", default="output/results.xlsx")
    parser.add_argument("--manual-review", default="output/manual_review.xlsx")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--reports-pdf", default="output/student_reports.pdf",
                         help="Output path for the single PDF containing every student's result")
    args = parser.parse_args()

    with open(args.config) as f:
        config = json.load(f)

    resolved_rolls = _read_manual_review(args.manual_review)
    if not resolved_rolls:
        print("No rows in manual_review.xlsx have an 'Actual Roll No' filled in yet.")
        return

    answer_key = _read_answer_key_from_report(args.results, config["marking_scheme"])
    photo_dir = config.get("student_photo_dir", "output/student_photos")

    wb = openpyxl.load_workbook(args.results)
    ws = wb["Results"]
    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header)}

    ws_detail = wb["Answer Detail"]
    detail_header = [c.value for c in ws_detail[1]]
    detail_file_col = detail_header.index("File") + 1
    detail_roll_col = 1  # "Roll No" is always column A

    done, still_invalid = 0, []

    for filename, actual_roll in resolved_rolls.items():
        student_info = fetch_student_info(actual_roll, config)
        if student_info is None:
            still_invalid.append((filename, actual_roll))
            continue
        student_info["photo_path"] = fetch_student_photo(
            student_info.get("image_url"), photo_dir, student_info["id_no"]
        )
        if student_info.get("image_url") and not student_info["photo_path"]:
            print(f"  Note: could not download photo for '{filename}' from {student_info['image_url']}")

        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=col["File"]).value == filename:
                ws.cell(row=r, column=col["Roll No"], value=actual_roll)
                ws.cell(row=r, column=col["ID No"], value=student_info["id_no"])
                ws.cell(row=r, column=col["Name"], value=student_info["name"])
                ws.cell(row=r, column=col["Father Name"], value=student_info["father_name"])
                ws.cell(row=r, column=col["Class"], value=student_info["class"])
                ws.cell(row=r, column=col["Section"], value=student_info["section"])
                ws.cell(row=r, column=col["Needs Manual Roll Review"], value="")
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = openpyxl.styles.PatternFill(fill_type=None)
                break

        for r in range(2, ws_detail.max_row + 1):
            if ws_detail.cell(row=r, column=detail_file_col).value == filename:
                ws_detail.cell(row=r, column=detail_roll_col, value=actual_roll)
                break

        done += 1
        print(f"Resolved '{filename}' -> roll {actual_roll} ({student_info['name']})")

    wb.save(args.results)

    # Rebuild the ONE combined result PDF from every currently-validated
    # row in Results (both the ones resolved just now and any that were
    # already validated in the original main.py run), since there's no
    # per-student file left on disk to incrementally add to.
    all_results = []
    for r in range(2, ws.max_row + 1):
        roll_no = ws.cell(row=r, column=col["Roll No"]).value
        if not roll_no or roll_no == "PENDING REVIEW":
            continue
        filename = ws.cell(row=r, column=col["File"]).value
        per_question = _reconstruct_per_question(ws_detail, filename, answer_key, config["marking_scheme"])
        if per_question is None:
            print(f"Could not find '{filename}' in the Answer Detail sheet - skipping from the combined PDF.")
            continue
        summary = _score_summary(per_question, config["marking_scheme"])
        id_no = ws.cell(row=r, column=col["ID No"]).value
        student_info = {
            "id_no": str(id_no) if id_no else "",
            "name": ws.cell(row=r, column=col["Name"]).value or "",
            "father_name": ws.cell(row=r, column=col["Father Name"]).value or "",
            "class": ws.cell(row=r, column=col["Class"]).value or "",
            "section": ws.cell(row=r, column=col["Section"]).value or "",
            "photo_path": _find_cached_photo(photo_dir, id_no),
        }
        all_results.append({
            "filename": filename, "roll_no": str(roll_no), "summary": summary,
            "per_question": per_question, "needs_review": False, "student_info": student_info,
        })

    combined_path, _ = build_combined_report(all_results, config, args.reports_pdf)

    print(f"\n{done} sheet(s) resolved.")
    if combined_path:
        print(f"Combined PDF rebuilt with all {len(all_results)} validated result(s): {combined_path}")
    if still_invalid:
        print(f"{len(still_invalid)} entered roll number(s) still not found in student records - re-check these:")
        for filename, roll in still_invalid:
            print(f"  {filename}: entered '{roll}'")


if __name__ == "__main__":
    main()
