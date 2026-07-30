"""
Generates a blank answer-key Excel template (sample_answer_key.xlsx)
with a row for every question (1-100) and empty/default columns for
the teacher to fill in before running main.py: Answer, Marks, Negative
Marks, Is Bonus, Is Multi Correct.

ANSWER column:
  - Most questions have one correct letter (e.g. "C").
  - "Is Multi Correct" = 0: multiple letters mean ALTERNATIVES - any
    ONE of them is accepted as the single correct answer (e.g. "B C"
    means "B or C", either alone is correct).
  - "Is Multi Correct" = 1: multiple letters must ALL be shaded
    together (e.g. "B C" means both B and C are required).
  - For anything more complex, write it out explicitly with and/or -
    this always overrides the Is Multi Correct column: "A and B",
    "C or D", or even "A and B or C and D".

MARKS / NEGATIVE MARKS: leave blank to use the batch-wide default from
config.json's marking_scheme; fill in to override for just that
question (e.g. a harder question worth more, or a question with no
negative marking).

IS BONUS: set to 1 to award every student full marks for that question
automatically, regardless of what they shaded (or left blank) - useful
for a question later found to be flawed/unanswerable.

Leave a question's Answer blank (and Is Bonus unset) if the exam
doesn't actually have that many questions - it will be excluded
entirely rather than counted against students.

Usage:
    python tools/generate_sample_key.py --num-questions 100 --output sample_answer_key.xlsx --config config.json
"""

import argparse
import json

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = ["Question", "Answer", "Marks", "Negative Marks", "Is Bonus", "Is Multi Correct"]


def generate(output_path, num_questions=100, default_marks=1.0, default_negative_marks=0.25):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AnswerKey"

    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for q in range(1, num_questions + 1):
        r = q + 1
        ws.cell(row=r, column=1, value=q)
        ws.cell(row=r, column=3, value=default_marks)
        ws.cell(row=r, column=4, value=default_negative_marks)
        ws.cell(row=r, column=5, value=0)
        ws.cell(row=r, column=6, value=0)

    # A free-text prompt, not a restrictive dropdown - multi-letter and
    # and/or expressions need to be typeable here.
    dv_answer = DataValidation(type="custom", formula1="TRUE", allow_blank=True)
    dv_answer.prompt = (
        'One letter (e.g. "C"). Multiple letters mean "any one of these" unless '
        'Is Multi Correct = 1 (then all must be shaded). For anything else, write '
        'it out: "A and B", "C or D", "A and B or C and D".'
    )
    dv_answer.promptTitle = "Correct answer(s)"
    ws.add_data_validation(dv_answer)
    dv_answer.add(f"B2:B{num_questions + 1}")

    dv_bool = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    dv_bool.prompt = "0 = No, 1 = Yes"
    ws.add_data_validation(dv_bool)
    dv_bool.add(f"E2:E{num_questions + 1}")
    dv_bool.add(f"F2:F{num_questions + 1}")

    dv_marks = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_marks.prompt = "Leave blank to use the batch-wide default from config.json"
    ws.add_data_validation(dv_marks)
    dv_marks.add(f"C2:C{num_questions + 1}")
    dv_marks.add(f"D2:D{num_questions + 1}")

    widths = [10, 26, 9, 14, 9, 15]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    wb.save(output_path)
    print(f"Saved blank answer key template ({num_questions} questions) to {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--num-questions", type=int, default=100)
    parser.add_argument("--output", default="sample_answer_key.xlsx")
    parser.add_argument("--config", default="config.json",
                         help="Read default Marks/Negative Marks from this config's marking_scheme, if present")
    args = parser.parse_args()

    default_marks, default_negative_marks = 1.0, 0.25
    try:
        with open(args.config) as f:
            scheme = json.load(f).get("marking_scheme", {})
        default_marks = scheme.get("correct", default_marks)
        default_negative_marks = abs(scheme.get("wrong", -default_negative_marks))
    except (FileNotFoundError, json.JSONDecodeError):
        pass

    generate(args.output, args.num_questions, default_marks, default_negative_marks)
